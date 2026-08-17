"""
Sincroniza a planilha de identidade dos grupos do OneDrive pro repositório.

O usuário edita a planilha direto no OneDrive (Base_Implantação/base
implantação.xlsx); este script copia a versão atual pra
data/base/base implantação.xlsx e dá commit + push sozinho, se algo mudou.
Feito pra rodar sem supervisão (Agendador de Tarefas do Windows) -- por
isso registra tudo em scripts/sync_planilha.log em vez de só print().

Rodar manualmente:
    python scripts/sync_planilha.py
"""

import filecmp
import logging
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ORIGEM = Path(r"C:\Users\warruda\OneDrive - Mgcontecnica\Base_Implantação\base implantação.xlsx")

RAIZ_REPO = Path(__file__).resolve().parent.parent
DESTINO = RAIZ_REPO / "data" / "base" / "base implantação.xlsx"
# Sidecar com a hora da última sincronização -- o backend expõe isso na API
# pro front mostrar "Base atualizada em ..." na barra superior. Guardado à
# parte (em vez de usar a data de modificação do arquivo) porque no deploy
# do Render o checkout do git muda a mtime do .xlsx pra hora do deploy, não
# pra hora real da última edição.
ARQUIVO_TIMESTAMP = DESTINO.parent / "atualizado_em.txt"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(Path(__file__).resolve().parent / "sync_planilha.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("sync_planilha")


def _ler_linhas(caminho):
    """Lê a aba 'Clientes' e devolve {ID: {coluna: valor}} -- ID é a
    chave de cada linha (empresa) na planilha, então serve pra comparar
    duas versões e achar o que mudou linha a linha."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Clientes"]
    cabecalho = [c.value for c in ws[1]]
    linhas = {}
    for linha in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(cabecalho, linha))
        id_empresa = d.get("ID")
        if id_empresa is None:
            continue
        linhas[id_empresa] = d
    return linhas


def _norm(valor):
    """None e string vazia contam como 'sem valor' -- evita ruído no diff
    (célula vazia vs. célula nunca preenchida não é uma mudança real)."""
    if valor is None:
        return ""
    return str(valor).strip()


def _nome_linha(d):
    return d.get("RazaoSocial") or d.get("Grupo") or "?"


def _log_diferencas(antigas, novas):
    """Compara as linhas antigas x novas por ID e loga o que mudou:
    empresas novas, removidas e, pras que continuam, quais campos
    mudaram de valor."""
    ids_antigos = set(antigas)
    ids_novos = set(novas)

    adicionadas = ids_novos - ids_antigos
    removidas = ids_antigos - ids_novos
    em_comum = ids_antigos & ids_novos

    alteracoes = {}
    for id_empresa in em_comum:
        antiga, nova = antigas[id_empresa], novas[id_empresa]
        campos = set(antiga) | set(nova)
        mudou = [
            (campo, antiga.get(campo), nova.get(campo))
            for campo in campos
            if campo != "ID" and _norm(antiga.get(campo)) != _norm(nova.get(campo))
        ]
        if mudou:
            alteracoes[id_empresa] = mudou

    if not (adicionadas or removidas or alteracoes):
        log.info("Planilha mudou (ex: formatação), mas nenhuma linha de dado foi alterada.")
        return

    log.info(
        "Alterações: %d empresa(s) nova(s), %d removida(s), %d alterada(s).",
        len(adicionadas), len(removidas), len(alteracoes),
    )
    for id_empresa in sorted(adicionadas, key=str):
        log.info("  + %s (ID %s)", _nome_linha(novas[id_empresa]), id_empresa)
    for id_empresa in sorted(removidas, key=str):
        log.info("  - %s (ID %s)", _nome_linha(antigas[id_empresa]), id_empresa)
    for id_empresa in sorted(alteracoes, key=str):
        nome = _nome_linha(novas[id_empresa])
        detalhes = "; ".join(
            f"{campo}: {_norm(antigo) or '(vazio)'} -> {_norm(novo) or '(vazio)'}"
            for campo, antigo, novo in alteracoes[id_empresa]
        )
        log.info("  * %s (ID %s): %s", nome, id_empresa, detalhes)


def _git(*args):
    """Roda um comando git na raiz do repo e devolve (codigo, stdout+stderr)."""
    resultado = subprocess.run(
        ["git", *args],
        cwd=RAIZ_REPO,
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    saida = (resultado.stdout + resultado.stderr).strip()
    return resultado.returncode, saida


def main():
    if not ORIGEM.exists():
        log.error("Planilha de origem não encontrada: %s (OneDrive sincronizado?)", ORIGEM)
        sys.exit(1)

    # Traz o repo em dia antes de mexer -- evita push rejeitado por
    # histórico divergente se alguém commitou outra coisa nesse meio tempo.
    codigo, saida = _git("pull", "--ff-only")
    if codigo != 0:
        log.error("git pull falhou, abortando sync:\n%s", saida)
        sys.exit(1)

    if DESTINO.exists() and filecmp.cmp(ORIGEM, DESTINO, shallow=False):
        log.info("Planilha já está igual, nada pra sincronizar.")
        return

    if DESTINO.exists():
        try:
            _log_diferencas(_ler_linhas(DESTINO), _ler_linhas(ORIGEM))
        except Exception:
            log.exception("Não deu pra calcular o diff da planilha (sync segue normalmente).")
    else:
        log.info("Primeira sincronização, sem versão anterior pra comparar.")

    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(ORIGEM, DESTINO)
    ARQUIVO_TIMESTAMP.write_text(datetime.now().astimezone().isoformat(), encoding="utf-8")
    log.info("Planilha copiada do OneDrive pro repo.")

    _git("add", str(DESTINO.relative_to(RAIZ_REPO)), str(ARQUIVO_TIMESTAMP.relative_to(RAIZ_REPO)))

    # git diff --cached --quiet devolve 0 se não há nada staged (pode
    # acontecer se o conteúdo binário mudou mas o git já tinha essa versão).
    codigo, _ = _git("diff", "--cached", "--quiet")
    if codigo == 0:
        log.info("Cópia igual à já commitada, nada pra commitar.")
        return

    codigo, saida = _git("commit", "-m", "Atualiza planilha de implantação (sync automático)")
    if codigo != 0:
        log.error("git commit falhou:\n%s", saida)
        sys.exit(1)

    codigo, saida = _git("push")
    if codigo != 0:
        log.error("git push falhou (commit ficou salvo localmente, tenta rodar de novo):\n%s", saida)
        sys.exit(1)

    log.info("Push feito com sucesso.")


if __name__ == "__main__":
    main()
