"""
Portal de Implantação — backend.

- Serve o site estático (index.html, static/, data/imagens).
- Identidade dos grupos (nome, contatos, equipe, dados contratuais,
  empresas, datas de implantação e se a implantação foi realizada) vem
  da planilha data/base/base implantação.xlsx (aba "Clientes"), só leitura pelo
  portal. Quem mantém essa informação em dia edita a planilha no OneDrive;
  um script separado (scripts/sync_planilha.py, rodando na máquina do
  usuário) copia a versão atualizada pra cá e dá commit/push sozinho.
  A coluna "GrupoID" é a chave que liga uma linha (empresa) da planilha
  ao grupo_id usado no Postgres; várias linhas com o mesmo GrupoID
  formam um grupo com várias empresas.
- Checklists (catálogo de itens + marcações) e observações — tudo que o
  portal de fato escreve — ficam no Postgres (variável DATABASE_URL).
  Schema em backend/schema_postgres.sql.

Rodar localmente:
    pip install -r backend/requirements.txt
    export DATABASE_URL=postgresql://usuario:senha@host/banco
    python backend/server.py
    abrir http://localhost:5000
"""

import html
import os
import re
from collections import OrderedDict
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel
import psycopg2
import psycopg2.pool
from psycopg2.extras import RealDictCursor
from flask import Flask, jsonify, request, send_from_directory

# ── Caminhos e config ────────────────────────────────────────
RAIZ = Path(__file__).resolve().parent.parent
ARQUIVO_EXCEL = RAIZ / "data" / "base" / "base implantação.xlsx"
# Escrito pelo scripts/sync_planilha.py a cada sync -- ver comentário lá
# sobre por que não usar a data de modificação do próprio .xlsx.
ARQUIVO_TIMESTAMP = RAIZ / "data" / "base" / "atualizado_em.txt"

DATABASE_URL = os.environ.get("DATABASE_URL")
if DATABASE_URL and DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

AREAS = ("dp", "ef", "ctb", "paralegal", "gerencia")

# Contato responsável por cada departamento: (sufixo das colunas na
# planilha -- Contato<sufixo>/Email<sufixo>/Telefone<sufixo> --, cargo
# mostrado no portal). Sufixo segue a mesma abreviação (CTB/EF/DP) e ordem
# já usadas nas colunas "Implantação <área>" da planilha; o cargo usa o
# nome como aparece no resto do portal (ex: tituloFiscal no app.js).
# Paralegal fica de fora -- decisão do usuário em 18/08/2026, não tem
# contato próprio por enquanto.
CONTATOS_POR_DEPARTAMENTO = (
    ("CTB", "Contábil"),
    ("EF", "Fiscal"),
    ("DP", "DP"),
)

app = Flask(__name__, static_folder=None)

# Pool de conexões: abrir uma conexão nova (handshake TLS com o Postgres do
# Render) a cada requisição era o principal motivo dos cards demorando pra
# abrir. Com o pool, a conexão é reaproveitada entre requisições.
_pool = None


def _obter_pool():
    global _pool
    if _pool is None:
        _pool = psycopg2.pool.ThreadedConnectionPool(
            1, 10, DATABASE_URL, cursor_factory=RealDictCursor
        )
    return _pool


@contextmanager
def get_conn():
    conn = _obter_pool().getconn()
    # autocommit evita que a conexão volte "suja" (transação aberta) pro
    # pool — sem isso, o putconn() faz um reset() com ida e volta extra ao
    # Postgres a cada requisição, quase dobrando o tempo de resposta.
    conn.autocommit = True
    try:
        yield conn
    finally:
        _obter_pool().putconn(conn)


# ── Identidade dos grupos (planilha Excel, só leitura) ─────────
def _texto(valor):
    """Limpa artefatos de exportação (';' e '/' sobrando no fim) e espaços."""
    if valor is None:
        return None
    s = re.sub(r"\s*/\s*$", "", str(valor).strip()).rstrip(";").strip()
    return s or None


def _para_datetime_excel(valor):
    """Converte o valor de uma célula de data em datetime, cobrindo tanto
    células formatadas como data (openpyxl já devolve datetime) quanto
    células digitadas sem formato de data aplicado -- nesse caso openpyxl
    devolve o número de série cru do Excel (ex: 46253) em vez de um
    datetime, e é isso que convertemos aqui."""
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, (int, float)) and valor > 0:
        try:
            return from_excel(valor)
        except (ValueError, OverflowError):
            return None
    return None


def _data_iso(valor):
    dt = _para_datetime_excel(valor)
    return dt.strftime("%Y-%m-%d") if dt else None


def _bool_sim_nao(valor):
    return str(valor or "").strip().lower() in ("sim", "true", "1", "x")


# Limpa um "(Sócio)" sobrando no fim do nome (convenção antiga da coluna
# Socio, hoje sem função -- ver carregar_identidade).
_RE_MARCA_SOCIO = re.compile(r"\s*[\(\[-]?\s*s[oó]cio\s*[\)\]]?\s*$", re.IGNORECASE)


_MAPA_ACENTOS = str.maketrans("áàãâéêíóõôúç", "aaaaeeiooouc")


def _slugificar(nome):
    """Deriva um grupo_id a partir do nome do grupo (ex: 'Loja Sul' ->
    'loja-sul'). Usado quando a planilha não preenche a coluna GrupoID —
    assim quem cadastra um grupo novo não precisa digitar o id na mão."""
    s = nome.strip().lower().translate(_MAPA_ACENTOS)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def carregar_identidade():
    """Lê a planilha e monta a mesma estrutura {"grupos": [...]} que a
    Tela 1/2 esperam — várias linhas (empresas) com o mesmo GrupoID
    viram um único grupo com várias "empresas". Se a coluna GrupoID
    estiver em branco numa linha, o id é derivado do nome do grupo."""
    wb = openpyxl.load_workbook(ARQUIVO_EXCEL, data_only=True)
    ws = wb["Clientes"]
    cabecalho = [c.value for c in ws[1]]

    grupos = OrderedDict()
    for linha in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(cabecalho, linha))
        nome_grupo = _texto(d.get("Grupo"))
        if not nome_grupo:
            continue
        grupo_id = _texto(d.get("GrupoID")) or _slugificar(nome_grupo)

        if grupo_id not in grupos:
            contrato_mg_dt = _para_datetime_excel(d.get("ContratoMG"))
            vigencia = contrato_mg_dt.strftime("%Y-%m") if contrato_mg_dt else None
            grupos[grupo_id] = {
                "id": grupo_id,
                "grupo": _texto(d.get("Grupo")),
                # "Realizada" agora é uma coluna por área na planilha (não existe
                # mais uma única "Implantação Realizada") -- cada área tem seu
                # próprio status porque uma pode estar concluída e outra não.
                # Paralegal não tem data/realizada na planilha (decisão do
                # usuário em 2026-08-10) -- fica de fora daqui de propósito.
                "implantacaoRealizada": {
                    "ctb": _bool_sim_nao(d.get("Implantação CTB - Realizada")),
                    "ef": _bool_sim_nao(d.get("Implantação EF - Realizada")),
                    "dp": _bool_sim_nao(d.get("Implantação DP - Realizada")),
                },
                "implantacao": {
                    "dp": _data_iso(d.get("Implantação DP")),
                    "ctb": _data_iso(d.get("Implantação CTB")),
                    "ef": _data_iso(d.get("Implantação EF")),
                    "paralegal": _data_iso(d.get("Implantação Paralegal")),
                },
                # nome -> contato; fica como dict até o fim da função pra
                # permitir trocar por um candidato mais específico (ver
                # comentário abaixo) sem perder a ordem de primeira aparição.
                "contatos": OrderedDict(),
                "equipe": {
                    "gerente": _texto(d.get("Gerente")),
                    "secretaria": _texto(d.get("Secretaria")),
                    "master": _texto(d.get("Master")),
                },
                "contratuais": {
                    "vigenciaContrato": vigencia,
                    "regimeContratual": _texto(d.get("Regime")),
                },
                "empresas": [],
            }
            # Contato por departamento é dado do grupo, não da empresa -- só
            # lido uma vez, na primeira linha do grupo (mesmo padrão de
            # "equipe" acima). Chave prefixada com "depto:" pra nunca colidir
            # com o dicionário de sócios/contatos (que usa o nome cru como
            # chave), mesmo se a mesma pessoa aparecer nos dois.
            for sufixo, cargo in CONTATOS_POR_DEPARTAMENTO:
                nome_contato = _texto(d.get(f"Contato {sufixo}"))
                if not nome_contato:
                    continue
                grupos[grupo_id]["contatos"][f"depto:{sufixo}"] = {
                    "nome": nome_contato,
                    "cargo": cargo,
                    "email": _texto(d.get(f"Email {sufixo}")),
                    "telefone": _texto(d.get(f"Telefone {sufixo}")),
                    "_especificidade": -1,
                }

        grupo = grupos[grupo_id]
        # html.unescape antes de separar: nomes colados de fontes já escapadas
        # em HTML (ex: "AA&amp;MARTINS") têm um ';' dentro da própria entidade
        # ("&amp;"), que quebraria o split abaixo no lugar errado.
        socio_texto = html.unescape(d.get("Socio") or "")
        nomes_brutos = [s.strip() for s in socio_texto.split(";") if s.strip()]
        # Todo nome na coluna Socio é sócio, ponto -- não precisa mais da
        # marca "(Sócio)" pra distinguir de contato direto, já que isso
        # agora tem colunas próprias (Contato CTB/EF/DP). Ainda limpamos um
        # "(Sócio)" sobrando no nome (convenção antiga, muita gente na
        # planilha ainda está marcada assim) só pra não repetir "Sócio" duas
        # vezes na tela -- não afeta mais o cargo, que já é sempre Sócio.
        pessoas = [
            {"nome": _RE_MARCA_SOCIO.sub("", nome_bruto).strip(), "cargo": "Sócio"}
            for nome_bruto in nomes_brutos
        ]

        # A planilha não tem e-mail/telefone por pessoa, só por empresa -- usa
        # o da linha (empresa) em que a pessoa apareceu. Quando a mesma pessoa
        # aparece em várias linhas (ex: numa linha-resumo do grupo com todos
        # juntos E na linha da empresa exclusiva dela), preferimos o contato
        # da linha mais específica -- menos nomes listados na mesma linha,
        # mais provável de ser o contato pessoal dela, em vez do contato
        # genérico do grupo repetido pra todo mundo.
        email_empresa = _texto(d.get("Email"))
        telefone_empresa = _texto(d.get("Telefone"))
        especificidade = len(pessoas)
        for pessoa in pessoas:
            atual = grupo["contatos"].get(pessoa["nome"])
            if atual is None:
                grupo["contatos"][pessoa["nome"]] = {
                    "nome": pessoa["nome"],
                    "cargo": pessoa["cargo"],
                    "email": email_empresa,
                    "telefone": telefone_empresa,
                    "_especificidade": especificidade,
                }
            elif especificidade < atual["_especificidade"]:
                atual["email"] = email_empresa
                atual["telefone"] = telefone_empresa
                atual["_especificidade"] = especificidade

        grupo["empresas"].append({
            "id": _texto(d.get("ID")),
            "razaoSocial": _texto(d.get("RazaoSocial")),
            "cnpj": None,
            "regime": _texto(d.get("Regime")),
            "segmento": _texto(d.get("Segmento")),
            "atividadeEconomica": None,
            "endereco": _texto(d.get("Endereco")),
            "email": _texto(d.get("Email")),
            "telefone": _texto(d.get("Telefone")),
            "socios": [p["nome"] for p in pessoas],
            "impostos": [],
        })

    for grupo in grupos.values():
        grupo["contatos"] = [
            {k: v for k, v in c.items() if k != "_especificidade"}
            for c in grupo["contatos"].values()
        ]

    return {"grupos": list(grupos.values())}


def encontrar_identidade(dados, grupo_id):
    return next((g for g in dados["grupos"] if g["id"] == grupo_id), None)


# ── Checklist (Postgres) ──────────────────────────────────────
def carregar_checklist(grupo_id):
    """Retorna {etapasPorArea, regimeEf, observacoes} para um grupo, ou
    None se o grupo ainda não existir em grupos_implantacao."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT regime_ef, observacoes FROM grupos_implantacao WHERE grupo_id = %s",
                (grupo_id,),
            )
            grupo_pg = cur.fetchone()
            if not grupo_pg:
                return None

            cur.execute("""
                SELECT i.id, i.departamento, i.texto, i.ordem, m.concluido
                FROM checklist_itens i
                JOIN checklist_marcacoes m ON m.checklist_item_id = i.id
                WHERE m.grupo_id = %s AND i.ativo
                ORDER BY i.departamento, i.ordem
            """, (grupo_id,))
            linhas = cur.fetchall()

    por_area = {a: [] for a in AREAS}
    for r in linhas:
        por_area[r["departamento"]].append({
            "id": r["id"],
            "nome": r["texto"],
            "concluido": r["concluido"],
            
        })

    return {
        "etapasPorArea": por_area,
        "regimeEf": grupo_pg["regime_ef"],
        "observacoes": grupo_pg["observacoes"],
    }


def carregar_progresso_geral_todos():
    """Progresso 'geral' de todos os grupos numa única query — evita abrir
    uma conexão nova por grupo (a Tela 1 lista todos de uma vez)."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT m.grupo_id,
                       count(*) AS total,
                       count(*) FILTER (WHERE m.concluido) AS feitos
                FROM checklist_marcacoes m
                JOIN checklist_itens i ON i.id = m.checklist_item_id AND i.ativo
                GROUP BY m.grupo_id
            """)
            linhas = cur.fetchall()
    return {
        r["grupo_id"]: round(r["feitos"] / r["total"] * 100) if r["total"] else 0
        for r in linhas
    }


def _regime_ef_padrao(segmento, regime_texto):
    """Decide automaticamente qual variante do checklist de Escrita Fiscal
    um grupo novo deve receber, a partir do que já vem na planilha:
    segmento Indústria sempre cai no checklist único de Fiscal Indústria;
    os demais (Varejo) seguem o regime tributário (texto tipo 'Federal - SN'
    vira Simples Nacional, o resto vira Lucro Real/Presumido)."""
    if "industria" in (segmento or "").lower():
        return "industria"
    if re.search(r"\bsn\b|simples", (regime_texto or "").lower()):
        return "simples_nacional"
    return "lucro_real_presumido"


def provisionar_grupos_novos(grupos):
    """Cria em grupos_implantacao (+ o checklist inicial em
    checklist_marcacoes) qualquer grupo que já exista na planilha mas
    ainda não tenha aparecido no Postgres — é assim que um grupo novo na
    planilha passa a ter checklist no portal sem precisar de SQL manual."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT grupo_id FROM grupos_implantacao")
            existentes = {r["grupo_id"] for r in cur.fetchall()}
            novos = [g for g in grupos if g["id"] not in existentes]
            if not novos:
                return

            cur.execute("SELECT id, departamento, regime FROM checklist_itens WHERE ativo")
            itens = cur.fetchall()

            for g in novos:
                empresas = g.get("empresas") or []
                segmento = empresas[0]["segmento"] if empresas else None
                regime_texto = empresas[0]["regime"] if empresas else None
                regime_ef = _regime_ef_padrao(segmento, regime_texto)

                cur.execute(
                    "INSERT INTO grupos_implantacao (grupo_id, regime_ef) VALUES (%s, %s) ON CONFLICT (grupo_id) DO NOTHING",
                    (g["id"], regime_ef),
                )
                for item in itens:
                    # item de EF só entra se for da variante certa pra esse grupo;
                    # itens dos outros departamentos (regime = NULL) entram sempre.
                    if item["departamento"] == "ef" and item["regime"] != regime_ef:
                        continue
                    cur.execute("""
                        INSERT INTO checklist_marcacoes (grupo_id, checklist_item_id, concluido)
                        VALUES (%s, %s, false)
                        ON CONFLICT (grupo_id, checklist_item_id) DO NOTHING
                    """, (g["id"], item["id"]))
            conn.commit()


def calcular_progresso(por_area):
    """% de conclusão geral e por área, a partir de etapasPorArea."""
    resultado = {"geral": 0, "dp": 0, "ef": 0, "ctb": 0, "paralegal": 0, "gerencia": 0}

    total_geral = 0
    feitos_geral = 0
    for area in AREAS:
        itens = por_area.get(area, [])
        total = len(itens)
        feitos = sum(1 for i in itens if i["concluido"])
        resultado[area] = round(feitos / total * 100) if total else 0
        total_geral += total
        feitos_geral += feitos

    resultado["geral"] = round(feitos_geral / total_geral * 100) if total_geral else 0
    return resultado


# ── API ──────────────────────────────────────────────────────
@app.get("/api/planilha-atualizada-em")
def planilha_atualizada_em():
    """Hora da última sincronização da planilha (scripts/sync_planilha.py),
    pra barra superior mostrar 'Base atualizada em ...'. None se o sync
    nunca rodou (ex: instalação nova sem o arquivo de timestamp ainda)."""
    if not ARQUIVO_TIMESTAMP.exists():
        return jsonify({"atualizadoEm": None})
    return jsonify({"atualizadoEm": ARQUIVO_TIMESTAMP.read_text(encoding="utf-8").strip()})


@app.get("/api/grupos")
def listar_grupos():
    """Lista enxuta para a Tela 1 (nome, datas por área e progresso)."""
    dados = carregar_identidade()
    provisionar_grupos_novos(dados["grupos"])
    progresso_por_grupo = carregar_progresso_geral_todos()
    lista = [
        {
            "id": g["id"],
            "grupo": g["grupo"],
            "implantacao": g.get("implantacao", {}),
            "implantacaoRealizada": g.get("implantacaoRealizada", {}),
            "vigencia": g.get("contratuais", {}).get("vigenciaContrato"),
            "progresso": progresso_por_grupo.get(g["id"], 0),
        }
        for g in dados["grupos"]
    ]
    lista.sort(key=lambda x: (x["vigencia"] or "9999-99", x["grupo"].lower()))
    return jsonify(lista)


@app.get("/api/grupos/<grupo_id>")
def obter_grupo(grupo_id):
    """Detalhe completo para a Tela 2: identidade (JSON) + checklist (Postgres)."""
    dados = carregar_identidade()
    grupo = encontrar_identidade(dados, grupo_id)
    if not grupo:
        return jsonify({"erro": "Grupo não encontrado"}), 404

    provisionar_grupos_novos([grupo])
    checklist = carregar_checklist(grupo_id)
    if not checklist:
        return jsonify({"erro": "Grupo sem checklist cadastrado no Postgres"}), 404

    resposta = dict(grupo)
    resposta["etapasPorArea"] = checklist["etapasPorArea"]
    resposta["regimeEf"] = checklist["regimeEf"]
    resposta["observacoes"] = checklist["observacoes"]
    resposta["progresso"] = calcular_progresso(checklist["etapasPorArea"])
    return jsonify(resposta)


@app.post("/api/grupos/<grupo_id>/etapas")
def marcar_etapa(grupo_id):
    """
    Marca/desmarca um item do checklist.
    Body JSON: { "etapaId": 5, "area": "ctb", "valor": true }
    """
    body = request.get_json(silent=True) or {}
    area = body.get("area")
    valor = bool(body.get("valor"))

    if area not in AREAS:
        return jsonify({"erro": "Área inválida (use dp, ef, ctb, paralegal ou gerencia)"}), 400

    try:
        etapa_id = int(body.get("etapaId"))
    except (TypeError, ValueError):
        return jsonify({"erro": "etapaId inválido"}), 400

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE checklist_marcacoes SET concluido = %s
                WHERE grupo_id = %s AND checklist_item_id = %s
                RETURNING id
            """, (valor, grupo_id, etapa_id))
            atualizado = cur.fetchone()
            conn.commit()

    if not atualizado:
        return jsonify({"erro": "Etapa não encontrada para esse grupo"}), 404

    checklist = carregar_checklist(grupo_id)
    return jsonify({"ok": True, "progresso": calcular_progresso(checklist["etapasPorArea"])})


@app.post("/api/grupos/<grupo_id>/observacoes")
def salvar_observacoes(grupo_id):
    """Salva o texto de observações do grupo. Body JSON: { "texto": "..." }"""
    body = request.get_json(silent=True) or {}
    texto = str(body.get("texto", ""))

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE grupos_implantacao SET observacoes = %s, atualizado_em = now()
                WHERE grupo_id = %s
                RETURNING grupo_id
            """, (texto, grupo_id))
            atualizado = cur.fetchone()
            conn.commit()

    if not atualizado:
        return jsonify({"erro": "Grupo não encontrado"}), 404
    return jsonify({"ok": True})


# ── Servir o site estático ───────────────────────────────────
@app.get("/")
def index():
    return send_from_directory(RAIZ, "index.html")


@app.get("/<path:caminho>")
def estaticos(caminho):
    # Serve static/, data/imagens, etc. Bloqueia acesso direto à planilha
    # (tem e-mail/telefone de cliente — não pode ser baixada pela URL).
    if caminho.replace("\\", "/").startswith("data/base/"):
        return jsonify({"erro": "Acesso negado"}), 403
    return send_from_directory(RAIZ, caminho)


if __name__ == "__main__":
    # use_reloader=False: o reloader do Flask atrapalha o pool de conexões
    # (cada request ficava lento como se abrisse conexão nova de novo).
    # Em produção (gunicorn) isso nem entra em jogo.
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)
